import pandas as pd
import numpy as np
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler

# We rely on the existing Connector class
from project_retail.project_retail.connectors.connector import Connector


def fetch_customer_features(conn: Connector) -> pd.DataFrame:
    """
    Return a DataFrame with columns: CustomerId, Age, Annual Income, Spending Score
    Data is joined from MySQL tables: customer and customer_spend_score
    """
    sql = (
        "select distinct customer.CustomerId, Age, Annual_Income, Spending_Score "
        "from customer, customer_spend_score "
        "where customer.CustomerId=customer_spend_score.CustomerID"
    )
    df = conn.queryDataset(sql)
    if df is None or df.empty:
        raise ValueError("No customer feature rows returned from MySQL.")
    # normalize column names
    df.columns = ['CustomerId', 'Age', 'Annual Income', 'Spending Score']
    return df


def kmeans_cluster(df_features: pd.DataFrame, feature_cols, n_clusters: int, scale: bool = False):
    """
    Run KMeans on selected feature columns.
    Returns a tuple (labels, model, X) where labels are numpy array of cluster labels.
    """
    X_raw = df_features.loc[:, feature_cols].values
    if scale:
        X = StandardScaler().fit_transform(X_raw)
    else:
        X = X_raw

    model = KMeans(n_clusters=n_clusters, init='k-means++', max_iter=500, random_state=42)
    labels = model.fit_predict(X)
    return labels.astype(int), model, X


def get_cluster_details(conn: Connector, df_features_with_labels: pd.DataFrame) -> dict:
    """
    For each cluster label in df_features_with_labels['cluster'], fetch full customer details
    from the MySQL 'customer' table.
    Returns dict: {cluster_label: DataFrame of full customer rows}
    """
    if 'cluster' not in df_features_with_labels.columns:
        raise ValueError("'cluster' column is required on features DataFrame")

    # Map cluster -> customer ids
    details = {}
    for cluster_label, sub in df_features_with_labels.groupby('cluster'):
        ids = sub['CustomerId'].tolist()
        if len(ids) == 0:
            details[cluster_label] = pd.DataFrame()
            continue
        # Build IN clause safely as integers
        ids = [int(i) for i in ids]
        id_list = ','.join(str(i) for i in ids)
        sql = f"select * from customer where CustomerId in ({id_list}) order by CustomerId"
        df_full = conn.queryDataset(sql)
        details[cluster_label] = df_full if df_full is not None else pd.DataFrame()
    return details


def print_customers_by_cluster(details: dict):
    """
    Pretty-print the full customer details grouped by cluster to console.
    """
    for cluster_label in sorted(details.keys()):
        df = details[cluster_label]
        count = 0 if df is None or df.empty else len(df)
        print(f"\n=== Cluster {cluster_label} ({count} customers) ===")
        if df is None or df.empty:
            print("(no rows)")
        else:
            print(df.to_string(index=False))


def export_clusters_to_excel(details: dict, output_path: str):
    """
    Export each cluster's customer details to a separate Excel sheet.
    Adds a Summary sheet with customer counts per cluster and applies simple formatting.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    # Build summary
    summary_rows = []
    for cluster_label in sorted(details.keys()):
        df = details[cluster_label]
        summary_rows.append({"cluster": cluster_label, "customer_count": 0 if df is None or df.empty else len(df)})
    summary_df = pd.DataFrame(summary_rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        for cluster_label, df in sorted(details.items()):
            sheet_name = f"Cluster_{cluster_label}"
            if df is None or df.empty:
                # write an empty sheet with a note
                pd.DataFrame({"note": ["no rows"]}).to_excel(writer, sheet_name=sheet_name, index=False)
                continue
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Simple formatting: auto width, freeze top row, auto filter for each sheet
        wb = writer.book
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row > 1:
                ws.freeze_panes = ws['A2']
                ws.auto_filter.ref = ws.dimensions
            # Auto column width
            widths = []
            for col_idx in range(1, ws.max_column + 1):
                max_len = 0
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    cell = row[0]
                    try:
                        val = str(cell.value) if cell.value is not None else ""
                    except Exception:
                        val = ""
                    max_len = max(max_len, len(val))
                widths.append(max_len + 2)
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = min(w, 40)